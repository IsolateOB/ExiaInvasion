import os
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def copy_sheet(ws_source, ws_target, start_row_target, min_row=1):
    row_offset = start_row_target - min_row
    max_row_source = ws_source.max_row
    max_col_source = ws_source.max_column

    for row in ws_source.iter_rows(min_row=min_row, max_row=max_row_source, max_col=max_col_source):
        for cell in row:
            new_cell = ws_target.cell(row=cell.row + row_offset, column=cell.column)
            copy_cell(cell, new_cell)

    for merged_range in ws_source.merged_cells.ranges:
        if merged_range.min_row >= min_row:
            start = merged_range.min_row + row_offset
            end = merged_range.max_row + row_offset
            new_range = f"{get_column_letter(merged_range.min_col)}{start}:" \
                        f"{get_column_letter(merged_range.max_col)}{end}"
            ws_target.merge_cells(new_range)

    for r in range(min_row, max_row_source + 1):
        if r in ws_source.row_dimensions:
            new_row = r + row_offset
            ws_target.row_dimensions[new_row].height = ws_source.row_dimensions[r].height

    if row_offset == 0:
        for col_letter, col_dim in ws_source.column_dimensions.items():
            if col_dim.width:
                ws_target.column_dimensions[col_letter].width = col_dim.width

    return max_row_source - min_row + 1

# 获取当前目录下所有 .xlsx 文件（排除合并后生成的文件）
files = sorted(glob.glob(os.path.join(os.getcwd(), "*.xlsx")))
if not files:
    print("未找到 .xlsx 文件！")
else:
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    current_row = 1

    for idx, file in enumerate(files):
        wb = load_workbook(file)
        ws = wb.active

        if idx == 0:
            rows_copied = copy_sheet(ws, merged_ws, current_row, min_row=1)
        else:
            rows_copied = copy_sheet(ws, merged_ws, current_row, min_row=4)
        current_row += rows_copied

    merged_ws.freeze_panes = merged_ws.cell(row=4, column=3)


    merged_wb.save("merged.xlsx")
    print("合并完成，生成文件：merged.xlsx")
