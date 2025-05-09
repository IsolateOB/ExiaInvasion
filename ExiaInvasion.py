from playwright.sync_api import sync_playwright
import json
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import chardet
import time
from http.cookies import SimpleCookie
import os


class ExiaInvasion:
    def __init__(self, language, browser, server, email, password, cookies):
        self.language = language
        self.browser = browser
        self.server = server
        self.email = email
        self.password = password
        self.cookies = cookies

        self.cube_dict_chs = {"遗迹突击魔方": {"cube_id": 1000301, "cube_level": 0},
                              "战术突击魔方": {"cube_id": 1000302, "cube_level": 0},
                              "遗迹巨熊魔方": {"cube_id": 1000303, "cube_level": 0},
                              "战术巨熊魔方": {"cube_id": 1000304, "cube_level": 0},
                              "遗迹促进魔方": {"cube_id": 1000305, "cube_level": 0},
                              "战术促进魔方": {"cube_id": 1000306, "cube_level": 0},
                              "遗迹量子魔方": {"cube_id": 1000307, "cube_level": 0},
                              "体力神器魔方": {"cube_id": 1000308, "cube_level": 0},
                              "遗迹强韧魔方": {"cube_id": 1000309, "cube_level": 0},
                              "遗迹治疗魔方": {"cube_id": 1000310, "cube_level": 0},
                              "遗迹回火魔方": {"cube_id": 1000311, "cube_level": 0},
                              "遗迹辅助魔方": {"cube_id": 1000312, "cube_level": 0},
                              "遗迹毁灭魔方": {"cube_id": 1000313, "cube_level": 0}}


        self.cube_dict_eng = {"Assault Cube": {"cube_id": 1000301, "cube_level": 0},
                             "Onslaught Cube": {"cube_id": 1000302, "cube_level": 0},
                             "Resilience Cube": {"cube_id": 1000303, "cube_level": 0},
                             "Bastion Cube": {"cube_id": 1000304, "cube_level": 0},
                             "Adjutant Cube": {"cube_id": 1000305, "cube_level": 0},
                             "Wingman Cube": {"cube_id": 1000306, "cube_level": 0},
                             "Quantum Cube": {"cube_id": 1000307, "cube_level": 0},
                             "Vigor Cube": {"cube_id": 1000308, "cube_level": 0},
                             "Endurance Cube": {"cube_id": 1000309, "cube_level": 0},
                             "Healing Cube": {"cube_id": 1000310, "cube_level": 0},
                             "Tempering Cube": {"cube_id": 1000311, "cube_level": 0},
                              "Relic Assist Cube": {"cube_id": 1000312, "cube_level": 0},
                              "Destruction Cube": {"cube_id": 1000313, "cube_level": 0}}


        self.cookie_str = self.get_cookies()
        if self.language == 1:
            self.account_dict = json.loads(open("SearchIndexEng.json", "r", encoding="utf-8").read())
        else:
            self.account_dict = json.loads(open("SearchIndexChs.json", "r", encoding="utf-8").read())
        self.account_dict["name"] = self.get_role_name()
        self.add_nikkes_details_to_dict()
        self.add_equipments_to_dict()
        self.save_dict_to_excel()


    def get_cookies(self):
        important_keys = ["OptanonAlertBoxClosed",
                          "game_login_game",
                          "game_openid",
                          "game_channelid",
                          "game_token",
                          "game_gameid",
                          "game_user_name",
                          "game_uid",
                          "game_adult_status",
                          "OptanonConsent"]

        if not pd.isna(self.cookies):
            cookie = SimpleCookie()
            cookie.load(self.cookies)

            parsed_cookie = {key: morsel.value for key, morsel in cookie.items()}
            filtered = {k: v for k, v in parsed_cookie.items() if k in important_keys}
            cookie_str = "; ".join([f"{k}={v}" for k, v in filtered.items()])
            if self.language == 1:
                print("Using existing cookies...")
            else:
                print("正在使用现有的cookie...")
            return cookie_str

        if self.language == 1:
            print("Please do not operate the browser unless there is human-machine verification, error reporting, or long-term inactivity, etc")
        else:
            print("请不要对浏览器进行任何操作，除非出现人机验证、报错、长时间无操作等情况")

        with sync_playwright() as p:
            if self.browser == 1:
                browser = p.chromium.launch(channel="msedge", headless=False)
            else:
                browser = p.chromium.launch(channel="chrome", headless=False)
            context = browser.new_context()
            page = context.new_page()
            page.goto("https://www.blablalink.com/login")


            page.click('#onetrust-accept-btn-handler')


            if self.server == 1:
                page.click(r"css=body > div.w-full.outline-none.max-h-\[65vh\].max-w-\[var\(--max-pc-w\)\].right-0.mx-auto.overflow-x-hidden.overflow-y-auto.flex.flex-col.bg-\[var\(--op-fill-white\)\].rounded-t-\[8px\].fixed.left-0.bottom-0.z-50 > div.flex-1.overflow-y-auto.w-full.mr-\[4px\].mb-\[35px\] > ul > li:nth-child(1)")
            else:
                page.click(r"css=body > div.w-full.outline-none.max-h-\[65vh\].max-w-\[var\(--max-pc-w\)\].right-0.mx-auto.overflow-x-hidden.overflow-y-auto.flex.flex-col.bg-\[var\(--op-fill-white\)\].rounded-t-\[8px\].fixed.left-0.bottom-0.z-50 > div.flex-1.overflow-y-auto.w-full.mr-\[4px\].mb-\[35px\] > ul > li:nth-child(2)")

            for i in range(20):
                if page.query_selector('#loginPwdForm_account'):
                    break
                time.sleep(0.1)

            if not page.query_selector('#loginPwdForm_account'):
                page.click('.pass-switchLogin__oper')

            page.fill('#loginPwdForm_account', self.email)
            page.fill('#loginPwdForm_password', self.password)
            page.click('xpath=//*[@id="loginPwdForm"]/div[3]/div/div/div/div/button')

            if self.language == 1:
                print("Retrieving cookies...")
            else:
                print("正在获取cookie...")

            for i in range(150):
                time.sleep(0.1)
                cookies = context.cookies()
                filtered = {c['name']: c['value'] for c in cookies if c['name'] in important_keys}
                if all(key in filtered for key in important_keys):
                    cookie_str = "; ".join([f"{k}={v}" for k, v in filtered.items()])
                    context.close()
                    browser.close()
                    return cookie_str

            if self.language == 1:
                print("Failed to retrieve cookies")
            else:
                print("获取cookie失败")

            context.close()
            browser.close()

            raise Exception("Failed to retrieve cookies")


    def get_header(self, content_length):
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "zh-CN,zh-TW;q=0.9,zh;q=0.8,en;q=0.7",
            "Content-Length": content_length,
            "Content-Type": "application/json",
            "Cookie": self.cookie_str,
            "Dnt": "1",
            "Origin": "https://www.blablalink.com",
            "priority": "u=1, i",
            "Referer": "https://www.blablalink.com/",
            "Sec-Ch-Ua": '"Chromium";v="134", "Not:A-Brand";v="24", "Microsoft Edge";v="134"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "sec-Fetch-Site": "same-site",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36 Edg/134.0.0.0",
            "x-channel-type": "2",
            "x-common-params": '{"game_id":"16","area_id":"global","source":"pc_web","intl_game_id":"29080","language":"zh-TW","env":"prod","data_statistics_scene":"outer","data_statistics_page_id":"https://www.blablalink.com/","data_statistics_client_type":"pc_web","data_statistics_lang":"zh-TW"}',
            "x-language": "zh-TW"
        }

        return headers


    def get_role_name(self):
        headers = self.get_header("2")
        url = "https://api.blablalink.com/api/ugc/direct/standalonesite/User/GetUserGamePlayerInfo"
        response = requests.post(url, headers=headers, json={}).json()

        role_name = response["data"]["role_name"]

        return role_name


    def get_player_nikkes(self):
        headers = self.get_header("2")
        url = "https://api.blablalink.com/api/game/proxy/Tools/GetPlayerNikkes"
        response = requests.post(url, headers=headers, json={}).json()

        return response


    def add_nikkes_details_to_dict(self):
        if self.language == 1:
            print("Fetching Nikke details...")
        else:
            print("正在获取Nikke详情...")

        player_nikkes = self.get_player_nikkes()

        for element, characters in self.account_dict["elements"].items():
            for character_name, details in characters.items():
                for nikke_details in player_nikkes["data"]["player_nikkes"]:
                    if details["name_code"] == nikke_details["name_code"]:
                        details["skill1_level"] = nikke_details["skill1_level"]
                        details["skill2_level"] = nikke_details["skill2_level"]
                        details["skill_burst_level"] = nikke_details["skill_burst_level"]
                        details["item_rare"] = nikke_details["item_rare"]
                        details["item_level"] = nikke_details["item_level"]
                        details["limit_break"] = nikke_details["limit_break"]
                    if nikke_details["level"] > self.account_dict["synchroLevel"]:
                        self.account_dict["synchroLevel"] = nikke_details["level"]
                    if self.language == 1:
                        for cube_name, cube_data in self.cube_dict_eng.items():
                            if nikke_details["cube_id"] == cube_data["cube_id"]:
                                if nikke_details["cube_level"] > cube_data["cube_level"]:
                                    self.cube_dict_eng[cube_name]["cube_level"] = nikke_details["cube_level"]
                    else:
                        for cube_name, cube_data in self.cube_dict_chs.items():
                            if nikke_details["cube_id"] == cube_data["cube_id"]:
                                if nikke_details["cube_level"] > cube_data["cube_level"]:
                                    self.cube_dict_chs[cube_name]["cube_level"] = nikke_details["cube_level"]



    def get_equipments(self, character_ids):
        headers = self.get_header("96")
        url = "https://api.blablalink.com/api/game/proxy/Tools/GetPlayerEquipContents"
        json_data = requests.post(url, headers=headers, json={"character_ids": character_ids}).json()
        if json_data is None:
            json_data = requests.post(url, headers=headers, json={"character_ids": character_ids}).json()
        player_equip_contents = json_data["data"]["player_equip_contents"]

        final_slots = [None, None, None, None]

        for record in reversed(player_equip_contents):
            equip_contents = record["equip_contents"]
            for i in range(4):
                if final_slots[i] is None:
                    slot_data = equip_contents[i]
                    if slot_data["equip_id"] != -99 or slot_data["equip_effects"]:
                        final_slots[i] = slot_data

        result = {}
        for slot_index, slot_data in enumerate(final_slots):
            if slot_data is None:
                result[slot_index] = []
                continue

            details_list = []
            for effect in slot_data["equip_effects"]:
                for func in effect["function_details"]:
                    details_list.append({
                        "function_type": func["function_type"],
                        "function_value": abs(func["function_value"]) / 100.0,
                        "level": func["level"]
                    })

            result[slot_index] = details_list

        return result


    def add_equipments_to_dict(self):
        if self.language == 1:
            print("Fetching equipment data...")
        else:
            print("正在获取装备数据...")

        for element, characters in self.account_dict["elements"].items():
            for character_name, details in characters.items():
                character_ids = [details["id"] + i for i in range(11)]
                details["equipments"] = self.get_equipments(character_ids)


    @staticmethod
    def set_outer_border(ws, start_row, start_col, end_row, end_col, side=Side(border_style="medium", color="000000")):
        for col in range(start_col, end_col + 1):
            top_cell = ws.cell(row=start_row, column=col)
            tborder = top_cell.border
            top_cell.border = Border(
                left=tborder.left,
                right=tborder.right,
                top=side,
                bottom=tborder.bottom
            )

            bottom_cell = ws.cell(row=end_row, column=col)
            bborder = bottom_cell.border
            bottom_cell.border = Border(
                left=bborder.left,
                right=bborder.right,
                top=bborder.top,
                bottom=side
            )

        for row in range(start_row, end_row + 1):
            left_cell = ws.cell(row=row, column=start_col)
            lborder = left_cell.border
            left_cell.border = Border(
                left=side,
                right=lborder.right,
                top=lborder.top,
                bottom=lborder.bottom
            )

            right_cell = ws.cell(row=row, column=end_col)
            rborder = right_cell.border
            right_cell.border = Border(
                left=rborder.left,
                right=side,
                top=rborder.top,
                bottom=rborder.bottom
            )


    @staticmethod
    def set_vertical_border(ws, start_row, end_row, col, border_side=Side(border_style="medium", color="000000"), side_pos="right"):
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=col)
            old_border = cell.border
            if side_pos == "right":
                cell.border = Border(
                    left=old_border.left,
                    right=border_side,
                    top=old_border.top,
                    bottom=old_border.bottom
                )
            else:
                cell.border = Border(
                    left=border_side,
                    right=old_border.right,
                    top=old_border.top,
                    bottom=old_border.bottom
                )


    @staticmethod
    def set_horizontal_border(ws, row, start_col, end_col, border_side=Side(border_style="medium", color="000000"), side_pos="bottom"):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=c)
            old_border = cell.border
            if side_pos == "bottom":
                cell.border = Border(
                    left=old_border.left,
                    right=old_border.right,
                    top=old_border.top,
                    bottom=border_side
                )
            else:
                cell.border = Border(
                    left=old_border.left,
                    right=old_border.right,
                    top=border_side,
                    bottom=old_border.bottom
                )


    @staticmethod
    def item_rare_to_str(item_rare):
        if item_rare == 1:
            return "R"
        elif item_rare == 2:
            return "SR"
        elif item_rare == 3:
            return "SSR"
        else:
            return ""


    @staticmethod
    def get_item_level(item_rare, item_level):
        if item_rare == 3:
            return f"{item_level + 1}★"
        else:
            return item_level



    @staticmethod
    def get_fill_by_level(level):
        if 1 <= level <= 5:
            return PatternFill("solid", fgColor="FF7777")  # 红
        elif 6 <= level <= 10:
            return PatternFill("solid", fgColor="FFFF77")  # 黄
        elif 11 <= level <= 14:
            return PatternFill("solid", fgColor="77AAFF")  # 蓝
        elif level == 15:
            return PatternFill("solid", fgColor="FF000000")  # 黑
        else:
            return None


    @staticmethod
    def get_font_by_level(level):
        if level == 15:
            return Font(color="FFFFFF")
        else:
            return Font(color="000000")


    @staticmethod
    def get_limit_break_str(limit_break):
        if limit_break < 0:
            limit_break_str = ""
        elif limit_break >= 0 and limit_break <= 3:
            limit_break_str = f"{limit_break} ★"
        elif limit_break > 3 and limit_break < 10:
            limit_break_str = f"+ {limit_break - 3}"
        else:
            limit_break_str = "MAX"
        return limit_break_str


    def save_dict_to_excel(self):
        if self.language == 1:
            print("Saving data to table...")
        else:
            print("正在保存数据到表格...")

        medium_side = Side(border_style="medium", color="FF000000")
        thin_side = Side(border_style="thin", color="FF000000")
        elements_data = self.account_dict["elements"]  # dict

        property_keys = [
            "limit_break",  # 0
            "skill1_level",  # 1
            "skill2_level",  # 2
            "skill_burst_level",  # 3
            "item_rare",  # 4
            "item_level",  # 5
            None,  # 6
            "IncElementDmg",    # 7
            "StatAtk",  # 8
            "StatAmmoLoad", # 9
            "StatChargeTime",   #10
            "StatChargeDamage",  # 11
            "StatCritical",  # 12
            "StatCriticalDamage",  # 13
            "StatAccuracyCircle",  # 14
            "StatDef", # 15
        ]
        property_labels_chs = [
            "突破",
            "技能1",
            "技能2",
            "爆裂",
            "珍藏品",  # 会合并到下一列
            None,  # 跳过
            "T10",
            "优越",
            "攻击",
            "弹夹",
            "蓄速",
            "蓄伤",
            "暴击",
            "暴伤",
            "命中",
            "防御"
        ]

        property_labels_eng = [
            "LB",
            "Skill 1",
            "Skill 2",
            "Burst",
            "Item",  # 会合并到下一列
            None,  # 跳过
            "T10",
            "Elem",
            "Atk",
            "Ammo",
            "Chg Spd",
            "Chg DMG",
            "Crit%",
            "Crit DMG",
            "Hit%",
            "Def"
        ]

        if self.language == 1:
            property_labels = property_labels_eng
        else:
            property_labels = property_labels_chs

        wb = Workbook()
        ws = wb.active
        if self.language == 1:
            ws.title = "Member Info"
        else:
            ws.title = "成员信息"

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 25

        # 表头
        if self.language == 1:
            cell_alliance = ws.cell(row=1, column=1, value="Name")
            cell_synchro = ws.cell(row=1, column=3, value="Synchro")
        else:
            cell_alliance = ws.cell(row=1, column=1, value="名称")
            cell_synchro = ws.cell(row=1, column=3, value="同步器")
        cell_alliance.font = Font(bold=True)
        cell_synchro.font = Font(bold=True)
        cell_alliance.alignment = Alignment(horizontal="center", vertical="center")
        cell_synchro.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=2)
        ws.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3)

        ws.merge_cells(start_row=4, start_column=1, end_row=8, end_column=1)  # 编号
        ws.merge_cells(start_row=4, start_column=2, end_row=8, end_column=2)  # 角色名称
        ws.merge_cells(start_row=4, start_column=3, end_row=8, end_column=3)  # 同步器

        # 在 row=4 写入角色名称和同步器的值
        ws.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=2, value=self.account_dict["name"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=2).font = Font(bold=True)

        ws.cell(row=4, column=3, value=self.account_dict["synchroLevel"]).alignment = Alignment(horizontal="center", vertical="center")

        start_col = 4
        width_per_char = 16

        for element_name, chars_dict in elements_data.items():
            num_chars = len(chars_dict)
            total_width = num_chars * width_per_char

            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + total_width - 1)
            c_elem = ws.cell(row=1, column=start_col, value=element_name)
            c_elem.alignment = Alignment(horizontal="center", vertical="center")
            c_elem.font = Font(bold=True)
            self.set_outer_border(ws, 1, start_col, 1, start_col + total_width - 1, medium_side)

            col_cursor = start_col
            for char_name, char_info in chars_dict.items():
                ws.merge_cells(start_row=2, start_column=col_cursor, end_row=2,
                               end_column=col_cursor + width_per_char - 1)
                c_char = ws.cell(row=2, column=col_cursor, value=char_name)
                c_char.alignment = Alignment(horizontal="center", vertical="center")
                self.set_horizontal_border(ws, 2, col_cursor, col_cursor + width_per_char - 1,
                                           border_side=thin_side, side_pos="bottom")

                priority = char_info.get("priority", "")
                if priority == "black":
                    c_char.fill = PatternFill("solid", fgColor="FF000000")
                    c_char.font = Font(color="FFFFFF", bold=True)
                elif priority == "blue":
                    c_char.fill = PatternFill("solid", fgColor="99CCFF")
                    c_char.font = Font(bold=True)
                elif priority == "yellow":
                    c_char.fill = PatternFill("solid", fgColor="FFFF88")
                    c_char.font = Font(bold=True)

                for i, label in enumerate(property_labels):
                    # 跳过 None
                    if label is None:
                        continue
                    col_index = col_cursor + i
                    if i == 4:
                        # 合并 i=4,5
                        ws.merge_cells(start_row=3, start_column=col_index, end_row=3, end_column=col_index + 1)
                        cell_head = ws.cell(row=3, column=col_index, value=label)
                        cell_head.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell_head = ws.cell(row=3, column=col_index, value=label)
                        cell_head.alignment = Alignment(horizontal="center", vertical="center")

                self.set_outer_border(ws, 2, col_cursor, 3, col_cursor + width_per_char - 1, medium_side)

                limit_break = char_info.get("limit_break", 0)
                skill1 = char_info.get("skill1_level", 0)
                skill2 = char_info.get("skill2_level", 0)
                skill_burst = char_info.get("skill_burst_level", 0)
                item_rare = char_info.get("item_rare", 0)
                item_rare_str = self.item_rare_to_str(item_rare)
                item_level = char_info.get("item_level", 0)

                limit_break_str = self.get_limit_break_str(limit_break)

                for i in range(6):
                    ws.merge_cells(
                        start_row=4, start_column=col_cursor + i,
                        end_row=8, end_column=col_cursor + i
                    )

                ws.cell(row=4, column=col_cursor, value=limit_break_str).alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(row=4, column=col_cursor + 1, value=skill1 if skill1 > 0 else "").alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(row=4, column=col_cursor + 2, value=skill2 if skill2 > 0 else "").alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(row=4, column=col_cursor + 3, value=skill_burst if skill_burst > 0 else "").alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(row=4, column=col_cursor + 4, value=item_rare_str).alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(row=4, column=col_cursor + 5, value=self.get_item_level(item_rare, item_level) if item_level >= 0 else "").alignment = Alignment(horizontal="center", vertical="center")

                if self.language == 1:
                    ws.cell(row=4, column=col_cursor + 6, value="Head").alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=5, column=col_cursor + 6, value="Body").alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=6, column=col_cursor + 6, value="Arm").alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=7, column=col_cursor + 6, value="Leg").alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=8, column=col_cursor + 6, value="Total").alignment = Alignment(horizontal="center", vertical="center")
                else:
                    ws.cell(row=4, column=col_cursor + 6, value="头").alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=5, column=col_cursor + 6, value="身").alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=6, column=col_cursor + 6, value="手").alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=7, column=col_cursor + 6, value="足").alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=8, column=col_cursor + 6, value="合计").alignment = Alignment(horizontal="center", vertical="center")

                equipments = char_info.get("equipments", {})
                sum_stats = {
                    "IncElementDmg": 0.0,
                    "StatAtk": 0.0,
                    "StatAmmoLoad": 0.0,
                    "StatChargeTime": 0.0,
                    "StatChargeDamage": 0.0,
                    "StatCritical": 0.0,
                    "StatCriticalDamage": 0.0,
                    "StatAccuracyCircle": 0.0,
                    "StatDef": 0.0,
                }
                for eq_idx in range(4):
                    row_idx = 4 + eq_idx
                    eq_list = equipments.get(eq_idx, [])
                    for i in range(7, 16):
                        c = ws.cell(row=row_idx, column=col_cursor + i)
                        c.value = ""
                        c.alignment = Alignment(horizontal="center", vertical="center")

                    for f in eq_list:
                        ftype = f.get("function_type", "")
                        fval = f.get("function_value", 0.0)
                        lvl = f.get("level", 0)
                        if ftype in sum_stats:
                            sum_stats[ftype] += fval
                        if ftype in property_keys[7:]:
                            i_prop = property_keys.index(ftype)  # 5..13
                            cell_eq = ws.cell(row=row_idx, column=col_cursor + i_prop)
                            cell_eq.value = fval/100
                            cell_eq.number_format = "0.00%"
                            # 上色
                            fill = self.get_fill_by_level(lvl)
                            font = self.get_font_by_level(lvl)
                            if fill: cell_eq.fill = fill
                            if font: cell_eq.font = font
                            cell_eq.alignment = Alignment(horizontal="center", vertical="center")

                for i in range(7, 16):
                    pkey = property_keys[i]
                    c_sum = ws.cell(row=8, column=col_cursor + i)
                    if pkey in sum_stats:
                        val_sum = sum_stats[pkey]
                        c_sum.value = val_sum/100
                        c_sum.number_format = "0.00%"
                    c_sum.alignment = Alignment(horizontal="center", vertical="center")

                self.set_outer_border(ws, 4, col_cursor, 8, col_cursor + width_per_char - 1, medium_side)

                self.set_vertical_border(ws, 3, 8, col_cursor, border_side=thin_side, side_pos="right")
                self.set_vertical_border(ws, 3, 8, col_cursor + 4, border_side=thin_side, side_pos="left")
                self.set_vertical_border(ws, 3, 8, col_cursor + 5, border_side=thin_side, side_pos="right")
                self.set_vertical_border(ws, 3, 8, col_cursor + 6, border_side=thin_side, side_pos="right")

                self.set_vertical_border(ws, 1, 8, 1, border_side=medium_side, side_pos="left")
                self.set_vertical_border(ws, 1, 8, 2, border_side=medium_side, side_pos="right")

                self.set_vertical_border(ws, 4, 8, 1, border_side=thin_side, side_pos="right")

                self.set_horizontal_border(ws, 3, 1, 3, side_pos="bottom")
                self.set_horizontal_border(ws, 8, 1, 3, side_pos="bottom")

                self.set_horizontal_border(ws, 8, col_cursor + 6, col_cursor + 15, border_side=thin_side, side_pos="top")

                col_cursor += width_per_char  # 下一个角色

            start_col += total_width

        if self.language == 1:
            ws.column_dimensions[get_column_letter(1)].width = 5
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 11

            for col in range(4, ws.max_column + 1):
                offset = (col - 4) % width_per_char
                if offset < 7:
                    ws.column_dimensions[get_column_letter(col)].width = 6
                else:
                    ws.column_dimensions[get_column_letter(col)].width = 10
        else:
            ws.column_dimensions[get_column_letter(1)].width = 5
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 8

            for col in range(4, ws.max_column + 1):
                offset = (col - 4) % width_per_char
                if offset < 6:
                    ws.column_dimensions[get_column_letter(col)].width = 6
                elif offset == 6:
                    ws.column_dimensions[get_column_letter(col)].width = 5
                else:
                    ws.column_dimensions[get_column_letter(col)].width = 10

        cube_start_col = col_cursor
        cube_count = len(self.cube_dict_chs)

        ws.merge_cells(start_row=1, start_column=cube_start_col, end_row=1, end_column=cube_start_col + cube_count - 1)
        if self.language == 1:
            cell_cube_header = ws.cell(row=1, column=cube_start_col, value="Cube")
        else:
            cell_cube_header = ws.cell(row=1, column=cube_start_col, value="魔方")
        cell_cube_header.alignment = Alignment(horizontal="center", vertical="center")
        cell_cube_header.font = Font(bold=True)
        self.set_outer_border(ws, 1, cube_start_col, 1, cube_start_col + cube_count - 1, medium_side)


        if self.language == 1:
            cube = self.cube_dict_eng
        else:
            cube = self.cube_dict_chs

        for i, (cube_name, cube_data) in enumerate(cube.items()):
            col = cube_start_col + i
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            cell_cube_name = ws.cell(row=2, column=col, value=cube_name)
            cell_cube_name.alignment = Alignment(horizontal="center", vertical="center")
            cell_cube_name.font = Font(bold=True)
            if i < cube_count - 1:
                self.set_vertical_border(ws, 2, 8, col, border_side=thin_side, side_pos="right")

        self.set_outer_border(ws, 2, cube_start_col, 3, cube_start_col + cube_count - 1, medium_side)

        for i, (cube_name, cube_data) in enumerate(cube.items()):
            col = cube_start_col + i
            ws.merge_cells(start_row=4, start_column=col, end_row=8, end_column=col)
            cube_level_value = cube_data["cube_level"]
            if cube_level_value == 0:
                if self.language == 1:
                    cube_level_value = "Not found"
                else:
                    cube_level_value = "未找到"
            cell_cube_level = ws.cell(row=4, column=col, value=cube_level_value)
            cell_cube_level.alignment = Alignment(horizontal="center", vertical="center")

        self.set_outer_border(ws, 4, cube_start_col, 8, cube_start_col + cube_count - 1, medium_side)


        if self.language == 1:
            for col in range(cube_start_col, cube_start_col + cube_count):
                ws.column_dimensions[get_column_letter(col)].width = 19
        else:
            for col in range(cube_start_col, cube_start_col + cube_count):
                ws.column_dimensions[get_column_letter(col)].width = 14


        new_font_name = "Microsoft YaHei"
        for row in ws.iter_rows():
            for cell in row:
                if cell.font:
                    old_font = cell.font
                    cell.font = Font(name=new_font_name,
                                     size=old_font.size,
                                     bold=old_font.bold,
                                     italic=old_font.italic,
                                     vertAlign=old_font.vertAlign,
                                     underline=old_font.underline,
                                     strike=old_font.strike,
                                     color=old_font.color)

        os.makedirs("output", exist_ok=True)
        filename = os.path.join("output", f"{self.account_dict['name']}.xlsx")
        wb.save(filename)

        if self.language == 1:
            print(f"Data saved to {filename}")
        else:
            print(f"数据已保存到 {filename}")


if __name__ == "__main__":
    print("ExiaInvasion v1.61  by 灵乌未默")
    print()
    print("GitHub:")
    print("github.com/IsolateOB/ExiaInvasion")
    print()

    print("1: English")
    print("2: 简体中文")
    print()

    print("Please select the language [1 or 2]:")
    print("请选择语言 [1或2]：")
    print()

    language = int(input())
    print()

    if language == 1:
        print("First run may not open the webpage properly and report errors continuously. Please close the program and browser and run again.")
        print()

        print("1: Edge")
        print("2: Chrome")
        print()

        browser = int(input("Please enter the browser number [1 or 2]:"))
        print()

        print("1: HK/MC/TW")
        print("2: JP/KR/NA/SEA/Global")
        print()

        server = int(input("Please enter the server number [1 or 2]:"))
        print()
    else:
        print("第一次运行可能无法正常打开网页并连续报错，请关闭程序与浏览器并重新运行")
        print()

        print("1: Edge")
        print("2: Chrome")
        print()

        browser = int(input("请输入浏览器编号[1或2]："))
        print()

        print("1: 香港/澳门/台湾")
        print("2: 日本/韩国/北美/东南亚/全球")
        print()

        server = int(input("请输入服务器编号[1或2]："))
        print()


    with open("LoginIndex.csv", "rb") as f:
        raw_data = f.read()
        encoding = chardet.detect(raw_data)["encoding"]

    loginIndex = pd.read_csv("LoginIndex.csv", encoding = encoding, dtype=str)


    def is_missing(val: str):
        if pd.isna(val):
            return True
        if isinstance(val, str):
            val_strip = val.strip()
            return val_strip == "" or val_strip.lower() == "nan"
        return False

    errorList = []
    total = len(loginIndex)
    i = 1
    for i, (idx, row) in enumerate(loginIndex.iterrows(), start=1):
        name = row.get("Name", "")
        email = row.get("E-mail", "")
        password = row.get("Password", "")
        cookies = row.get("Cookies", "")

        if is_missing(cookies):
            if is_missing(email) or is_missing(password):
                errorList.append((i, name))
                if language == 1:
                    print(f"Skip ({i}/{total}) {name} due to missing information")
                else:
                    print(f"缺失信息，跳过 ({i}/{total}) {name}")
                continue

        if language == 1:
            print(f"Logging in with account ({i}/{len(loginIndex)}): {name}")
        else:
            print(f"正在登录账号 ({i}/{len(loginIndex)}): {name}")
        try:
            ExiaInvasion(language, browser, server, email, password, cookies)
        except Exception:
            if language == 1:
                print(f"Error occurred while processing account {i}: {name}")
            else:
                print(f"处理账号 {i} 时发生错误: {name}")
            errorList.append((i, name))

        print()


    error_count = len(errorList)

    if language == 1:
        print(f"All accounts processed. Total errors: {error_count}")
    else:
        print(f"所有账号处理完成。总错误数: {error_count}")
    print()

    if error_count > 0:
        if language == 1:
            print("Error accounts:")
        else:
            print("错误账号：")

        with open("ErrorList.txt", "w", encoding="utf-8") as f:
            for error in errorList:
                print(f"{error[0]}: {error[1]}")
                f.write(f"{error[0]}: {error[1]}\n")
            print()

        if language == 1:
            print("Error account list generated: ErrorList.txt")
        else:
            print("已生成错误账号清单：ErrorList.txt")

    if language == 1:
        input("Press Enter to exit...")
    else:
        input("按回车键退出...")
